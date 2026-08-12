import openpyxl

file_path = r"C:\Users\riyal\Documents\RCR-ERP\NEO ITURKAA ENTERPRISES RA BILL 20 MAY 26.xlsx"

try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    print("Sheets in the workbook:")
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]
        # print first 30 rows and columns up to 20 to understand structure
        for row in sheet.iter_rows(min_row=1, max_row=30, min_col=1, max_col=15):
            row_data = []
            for cell in row:
                val = cell.value
                if val is not None:
                    row_data.append(f"{cell.coordinate}: {val}")
            if row_data:
                print(" | ".join(row_data))
except Exception as e:
    print(f"Error reading excel: {e}")
